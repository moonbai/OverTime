# core/data_manager.py
import os
import csv
import shutil
from datetime import datetime
from typing import List, Tuple, Dict, Any

class DataManager:
    """数据管理器"""

    def __init__(self, data_dir="data"):
        self.data_dir = data_dir
        self.overtime_file = os.path.join(data_dir, "overtime_records.csv")
        self.backup_dir = os.path.join(data_dir, "backup")

        # 确保data目录存在
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            print(f"✓ 创建数据目录: {self.data_dir}")

        # 确保backup目录存在
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
            print(f"✓ 创建备份目录: {self.backup_dir}")

    def create_file_if_not_exists(self):
        """如果文件不存在则创建CSV文件 - 使用UTF-8-sig"""
        if not os.path.exists(self.overtime_file):
            # 🎯 使用 utf-8-sig 编码（带BOM，Excel可识别）
            with open(self.overtime_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["日期", "用户", "类型", "加班时长", "请假类型", "请假时长", "提交时间", "加班工资"])
            print(f"✓ 创建数据文件: {self.overtime_file}")
        else:
            print(f"ℹ CSV文件已存在: {self.overtime_file}")

    def add_record(self, record: List[str]) -> bool:
        """添加记录 - 使用UTF-8-sig"""
        try:
            # 🎯 使用 utf-8-sig 编码
            with open(self.overtime_file, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(record)
            return True
        except Exception as e:
            print(f"✗ 添加记录失败: {e}")
            return False

    def get_all_records(self) -> List[List[str]]:
        """获取所有记录"""
        if not os.path.exists(self.overtime_file):
            return []

        # 🎯 尝试多种编码
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']
        for encoding in encodings:
            try:
                with open(self.overtime_file, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    next(reader, None)  # 跳过表头
                    records = [row for row in reader if row]
                    if records:
                        print(f"✓ 使用编码 {encoding} 读取成功")
                        return records
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"⚠ 编码 {encoding} 失败: {e}")
                continue

        print("✗ 所有编码尝试失败")
        return []

    def get_all_records_with_total(self) -> Tuple[List[List[str]], int]:
        """获取所有记录和总数"""
        records = self.get_all_records()
        return records, len(records)

    def get_monthly_records(self, month: str) -> List[List[str]]:
        """获取某月记录"""
        all_records = self.get_all_records()
        return [r for r in all_records if r and r[0].startswith(month)]

    def get_filtered_records(self, filters: Dict[str, Any]) -> Tuple[List[List[str]], int]:
        """获取筛选后的记录"""
        all_records = self.get_all_records()
        filtered = all_records

        # 用户筛选
        if 'user' in filters:
            filtered = [r for r in filtered if filters['user'] in r[1]]

        # 日期筛选
        if 'date_start' in filters:
            filtered = [r for r in filtered if r[0] >= filters['date_start']]
        if 'date_end' in filters:
            filtered = [r for r in filtered if r[0] <= filters['date_end']]

        # 类型筛选
        if 'type' in filters:
            filtered = [r for r in filtered if r[2] == filters['type']]

        return filtered, len(filtered)

    def import_csv(self, file_path: str, default_user: str = "未知") -> Tuple[int, int, List[str]]:
        """导入CSV记录"""
        imported = 0
        failed = 0
        errors = []

        # 🎯 尝试多种编码
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'big5']

        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    headers = next(reader, None)

                    for i, row in enumerate(reader, 1):
                        if not row:
                            continue

                        try:
                            # 处理可能的空字段
                            if len(row) < 8:
                                row.extend([""] * (8 - len(row)))

                            # 确保有用户
                            if not row[1]:
                                row[1] = default_user

                            # 验证日期格式
                            datetime.strptime(row[0], "%Y-%m-%d")

                            if self.add_record(row):
                                imported += 1
                            else:
                                failed += 1
                                errors.append(f"第{i}行: 保存失败")
                        except ValueError:
                            failed += 1
                            errors.append(f"第{i}行: 日期格式错误")
                        except Exception as e:
                            failed += 1
                            errors.append(f"第{i}行: {str(e)}")

                        if len(errors) >= 5:
                            errors.append("...更多错误省略")
                            break

                print(f"✓ 使用编码 {encoding} 导入成功")
                return imported, failed, errors[:5]

            except UnicodeDecodeError:
                continue
            except Exception as e:
                errors.append(f"编码 {encoding} 失败: {str(e)}")
                continue

        return 0, 0, ["文件读取失败，所有编码尝试均失败"]

    def export_excel(self, file_path: str) -> bool:
        """导出到Excel"""
        try:
            # 检查openpyxl是否安装
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                print("✗ 未安装openpyxl，请执行: pip install openpyxl")
                return False

            records = self.get_all_records()

            if not records:
                print("⚠ 没有数据可导出")
                return False

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "加班记录"

            # 表头
            headers = ["日期", "用户", "类型", "加班时长", "请假类型", "请假时长", "提交时间", "加班工资"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font.color = "FFFFFF"
                cell.alignment = Alignment(horizontal="center")

            # 数据
            for row_idx, record in enumerate(records, 2):
                for col_idx, value in enumerate(record, 1):
                    # 🎯 确保值不为None
                    cell_value = str(value) if value is not None else ""
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell_value = str(cell.value) if cell.value else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                # 最小宽度8，最大宽度50
                ws.column_dimensions[column].width = min(max(max_length + 2, 8), 50)

            #确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # 保存文件
            wb.save(file_path)
            print(f"✓ Excel文件已保存: {file_path}")
            return True

        except PermissionError:
            print(f"✗ 权限错误，文件可能被占用: {file_path}")
            return False
        except Exception as e:
            print(f"✗ 导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def backup(self) -> bool:
        """备份数据文件"""
        if not os.path.exists(self.overtime_file):
            return False

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(self.backup_dir, f"overtime_records_{timestamp}.csv")
            shutil.copy2(self.overtime_file, backup_file)
            print(f"✓ 备份成功: {backup_file}")
            return True
        except Exception as e:
            print(f"✗ 备份失败: {e}")
            return False
